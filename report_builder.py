"""Excel report builder — produces a professional forensic report using openpyxl.
Runs anywhere, no Excel installation required.

Structure:
  1. Cover — client name, period, date of report
  2. Summary — all tests with exception counts
  3. Assumptions — country, weekend days, holidays used, account classification
  4. T01..T14 — one tab per test with findings table and conclusion
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
PINK = PatternFill("solid", start_color="FFC8FF")
LIGHT_BLUE = PatternFill("solid", start_color="DDEBF7")
DARK_BLUE = PatternFill("solid", start_color="1F4E78")
GREY = PatternFill("solid", start_color="F2F2F2")
RED_TINT = PatternFill("solid", start_color="FCE4E4")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style(cell, bold=False, size=10, color="000000", fill=None, align=None, border=False):
    cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = BORDER


def _write_header(ws, client_name, client_period, subject):
    _style(ws["A1"], bold=True, size=14); ws["A1"] = client_name
    _style(ws["A2"], bold=True, size=11); ws["A2"] = client_period
    _style(ws["A3"], bold=True, size=11); ws["A3"] = subject


def _cover(wb, client_name, client_period, results):
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    ws.merge_cells("A1:B1")
    _style(ws["A1"], bold=True, size=22, color="FFFFFF", fill=DARK_BLUE, align="center")
    ws["A1"] = "Journal Entry Testing Report"
    ws.row_dimensions[1].height = 40

    rows = [
        ("Client", client_name),
        ("Period covered", client_period),
        ("Report date", datetime.now().strftime("%d %B %Y")),
        ("Total tests run", sum(1 for r in results if not r.skipped)),
        ("Tests skipped", sum(1 for r in results if r.skipped)),
        ("Total exceptions flagged", sum(r.count for r in results if not r.skipped)),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        _style(ws[f"A{i}"], bold=True, size=11, fill=GREY, border=True); ws[f"A{i}"] = label
        _style(ws[f"B{i}"], size=11, border=True); ws[f"B{i}"] = val

    ws[f"A{len(rows) + 5}"] = "This report was produced by Ledger Scanner. Findings require auditor review — not all flagged entries are errors or fraud."
    _style(ws[f"A{len(rows) + 5}"], size=9, color="808080")
    ws.merge_cells(f"A{len(rows) + 5}:B{len(rows) + 5}")


def _summary(wb, results):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15

    headers = ("Test ID", "Test", "Status", "Exceptions")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        _style(c, bold=True, fill=PINK, align="center", border=True)

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r.test_id)
        ws.cell(row=i, column=2, value=r.title)
        if r.skipped:
            ws.cell(row=i, column=3, value="Skipped")
            ws.cell(row=i, column=4, value="—")
        else:
            ws.cell(row=i, column=3, value="Run")
            ws.cell(row=i, column=4, value=r.count)

        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            _style(cell, size=10, border=True)
            if r.skipped:
                cell.fill = GREY
            elif r.count > 0:
                cell.fill = RED_TINT


def _assumptions(wb, ctx, cols):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    _style(ws["A1"], bold=True, size=13, fill=DARK_BLUE, color="FFFFFF", align="center")
    _style(ws["B1"], fill=DARK_BLUE)
    ws.merge_cells("A1:B1")
    ws["A1"] = "Engagement configuration & assumptions"

    entries = [
        ("Country / jurisdiction", ctx.get("country_name", "—")),
        ("Years covered (for holidays)", ", ".join(str(y) for y in ctx.get("years", []))),
        ("Weekend days", ctx.get("weekend_label", "—")),
        ("Public holidays identified", f"{len(ctx.get('holidays', []))} dates"),
        ("Custom holidays added", f"{len(ctx.get('custom_holidays', []))} dates"),
        ("Odd-hours window", f"{ctx.get('odd_hours_start', 19)}:00 to {ctx.get('odd_hours_end', 8)}:00"),
        ("Bank accounts classified", len(ctx.get("bank_codes", []))),
        ("Prepayment accounts classified", len(ctx.get("prepayment_codes", []))),
        ("Accrual accounts classified", len(ctx.get("accrual_codes", []))),
        ("Revenue accounts classified", len(ctx.get("revenue_codes", []))),
        ("P&L accounts classified", len(ctx.get("pnl_codes", []))),
        ("Key accounts classified", len(ctx.get("key_account_codes", []))),
        ("Suspicious keywords used", len(ctx.get("suspicious_words", []))),
    ]
    for i, (k, v) in enumerate(entries, start=3):
        _style(ws.cell(row=i, column=1, value=k), bold=True, fill=GREY, border=True)
        _style(ws.cell(row=i, column=2, value=v), border=True)

    # Column mapping block
    start = len(entries) + 5
    _style(ws.cell(row=start, column=1, value="Column mapping used"), bold=True, size=12, fill=DARK_BLUE, color="FFFFFF")
    _style(ws.cell(row=start, column=2), fill=DARK_BLUE)
    ws.merge_cells(start_row=start, end_row=start, start_column=1, end_column=2)
    for j, (field, col) in enumerate(cols.items(), start=start + 1):
        _style(ws.cell(row=j, column=1, value=field.replace("_", " ").title()), bold=True, fill=GREY, border=True)
        _style(ws.cell(row=j, column=2, value=col if col else "— not mapped —"), border=True)


def _test_tab(wb, result, client_name, client_period, idx):
    sheet_name = f"T{idx:02d}"[:31]
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, client_name, client_period, result.title)

    _style(ws["A5"], bold=True, size=10); ws["A5"] = f"Objective: {result.objective}"
    _style(ws["A6"], size=10); ws["A6"] = f"Method: {result.method}"
    ws.row_dimensions[5].height = 30; ws.row_dimensions[6].height = 30
    ws.merge_cells("A5:H5"); ws.merge_cells("A6:H6")

    if result.skipped:
        _style(ws["A8"], bold=True, size=11, fill=GREY); ws["A8"] = "Test skipped"
        _style(ws["A9"], size=10); ws["A9"] = result.notes or "Required inputs not available."
        return

    if result.flagged is None or result.flagged.empty:
        _style(ws["A8"], bold=True, size=11); ws["A8"] = result.conclusion
        return

    headers = list(result.flagged.columns)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=str(h))
        _style(cell, bold=True, fill=PINK, align="center", border=True)

    for row_idx, (_, row) in enumerate(result.flagged.iterrows(), start=9):
        for col_idx, val in enumerate(row, start=1):
            display = "" if (val is None or (isinstance(val, float) and val != val)) else val
            # Convert Timestamps to date strings for Excel
            if hasattr(display, "strftime"):
                display = display.strftime("%Y-%m-%d") if hasattr(display, "date") else str(display)
            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            _style(cell, size=10, border=True)
            if row_idx % 2 == 0:
                cell.fill = LIGHT_BLUE

    # Auto-width (capped)
    for col_idx, h in enumerate(headers, start=1):
        sample = result.flagged[h].astype(str).head(200)
        max_len = max([len(str(h))] + [len(s) for s in sample if s])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

    concl_row = 9 + len(result.flagged) + 2
    _style(ws.cell(row=concl_row, column=1, value=f"Conclusion: {result.conclusion}"), bold=True, size=11)
    ws.merge_cells(start_row=concl_row, end_row=concl_row, start_column=1, end_column=min(len(headers), 8))


def build_report(client_name, client_period, results, ctx=None, cols=None):
    """Returns bytes of the complete .xlsx report."""
    ctx = ctx or {}
    cols = cols or {}

    wb = Workbook()
    _cover(wb, client_name, client_period, results)
    _summary(wb, results)
    _assumptions(wb, ctx, cols)

    for i, r in enumerate(results, start=1):
        _test_tab(wb, r, client_name, client_period, i)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
